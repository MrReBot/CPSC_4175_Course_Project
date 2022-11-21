import os
from PyPDF2 import PdfFileReader, PdfFileWriter
import re
import openpyxl as xl
import  Database

def parse_file(filename:str, db) -> list:
    """Get a files extension and attempt to parse it into a class list"""
    ext =  filename[filename.rfind(".")+1:] # Get the file extension
    if not os.path.exists(filename):
        print("ERROR: File Doesn't exist!")
        return
    if ext == "txt":
        return parse_txt(filename, db)
    elif ext == "pdf":
        return parse_pdf(filename, db)
    else:
        print(f"'{ext}' isn't a supported file extension")



def parse_txt(filename: str, db) -> list:
    """Parser for text files"""
    course_list = []
    with open(filename,"r") as f:
        for line in f.read().splitlines():
            if not line.startswith("#"):
                course_list.append(line.strip())
        return course_list

def parse_pdf(filename: str, db) -> list:
    """Parser for DegreeWorks PDF Files"""
    pdf = PdfFileReader(filename)
    course_list = []
    prune_list = ["CPSC 5115"] + db.get_tag("SCIE-LABB")
    for page_num in range(pdf.numPages):
        pageObj = pdf.getPage(page_num)
        txt = pageObj.extract_text().splitlines()
        for line_num, line in enumerate(txt):
            if "SCIENCE COURSES WITH LAB Still Needed" in line:
                count = int(re.findall("[0-9]+", line)[0]) #Get the number of labs that are required
                for i in range(count):
                     course_list.append("SCIE-LABB")
            #elif "Senior Software Engineering Project" in line:
            #    course_list.append("CPSC 4176")
            elif "Class in" in line:
                course = re.findall("[a-zA-Z]+ [0-9]{4}",line)
                try:
                    course_list.append(course[0])
                except IndexError:
                    pass
            #elif "Electives" in line: # Need this for elective processing
            #    elective = db.get_elective(line.split(" Electives ")[0])
            #    credits = txt[line_num+1].strip().split(" ")[0]
            #    course_list.append(f"ELECT,{elective},{credits}")
            elif "Credits in" in line:
                continue
                line = line.strip() #Remove any trailing or leading whitespace
                elective = []
                credits = line.split(" ")[0]
                if credits.isnumeric():
                    count = credits // 3 # Calculate how many classes this is
                else:
                    continue
                for course in re.findall("[a-zA-Z]{4} [0-9]@", line):
                    course = course.split(" ")[0] #Remove the 3@
                    elective.append(course)
                for i in range(count):
                    course_list.append("/".join(elective)+"-ELECT")
            
    for course in prune_list:
        if course in course_list:
            course_list.remove(course)
    return course_list

def excelParser(self, excelFileName):
    #If file does not exists then print an error and return else parse the file into useable data
    if(os.path.exists(excelFileName)==False):
        print("ERROR: File Doesn't exist!")
        return
    #else load the file and run it through the dictionary creator
    else:
        wb = xl.load_workbook(excelFileName)
        ws =wb.active           
        classdict = dictCreator(self, ws)[0]
        classlist = dictCreator(self, ws)[1]
        #retutn the created dictionary
        return classdict, classlist
        
def dictCreator(self, ws):
     #variables required in order to create a sorted dictionary from an excel sheet
     excelindex = 0
     classdict = {}
     courselist =[]
     FallSemesterName = ""
     SpringSemesterName = ""
     SummerSemesterName = ""
     #The while loop runs and sorts the data based on the semester name and the class information
     while excelindex <36:
         if(ws['A'+str(excelindex+3)].value !=None):
             #If the data does not contain the word fall or total, add it to the list under the existing fall semester key
             if('Fall' not in ws['A'+str(excelindex+3)].value and 'Total' not in ws['A'+str(excelindex+3)].value ):
                 classdict[FallSemesterName].append(ws['A'+str(excelindex+3)].value)
                 classdict[FallSemesterName].append(ws['B'+str(excelindex+3)].value)
                 courselist.append(ws['A'+str(excelindex+3)].value)
             #else if the data is a semester name, change the fallsemestername to that semester, and add the key to the dictionary
             elif ('Fall' in ws['A'+str(excelindex+3)].value and "__" not in ws['A'+str(excelindex+3)].value):
                 classdict[ws['A'+str(excelindex+3)].value] = list()
                 FallSemesterName = ws['A'+str(excelindex+3)].value
         
         if(ws['C'+str(excelindex+3)].value !=None):
             #If the data does not contain the word spring or total, add it to the list under the existing spring semester key
             if('Spring' not in ws['C'+str(excelindex+3)].value and 'Total' not in ws['C'+str(excelindex+3)].value ):
                 classdict[SpringSemesterName].append(ws['C'+str(excelindex+3)].value)
                 classdict[SpringSemesterName].append(ws['D'+str(excelindex+3)].value)
                 courselist.append(ws['C'+str(excelindex+3)].value)
             #else if the data is a semester name, change the springsemestername to that semester, and add the key to the dictionary
             elif ('Spring' in ws['C'+str(excelindex+3)].value and "__" not in ws['C'+str(excelindex+3)].value):
                 classdict[ws['C'+str(excelindex+3)].value] = list()
                 SpringSemesterName = ws['C'+str(excelindex+3)].value
                 
         if(ws['E'+str(excelindex+3)].value !=None):
             #If the data does not contain the word summer or total, add it to the list under the existing summer semester key
             if('Summer' not in ws['E'+str(excelindex+3)].value and 'Total' not in ws['E'+str(excelindex+3)].value ):
                 classdict[SummerSemesterName].append(ws['E'+str(excelindex+3)].value)
                 classdict[SummerSemesterName].append(ws['F'+str(excelindex+3)].value)
                 courselist.append(ws['E'+str(excelindex+3)].value)
                 #else if the data is a semester name, change the summersemestername to that semester, and add the key to the dictionary
             elif ('Summer' in ws['E'+str(excelindex+3)].value and "__" not in ws['E'+str(excelindex+3)].value):
                 classdict[ws['E'+str(excelindex+3)].value] = list()
                 SummerSemesterName = ws['E'+str(excelindex+3)].value
                 
         excelindex+=1
     return classdict, courselist

if __name__ == "__main__":
    db = Database.Database("database.txt")
    print(parse_pdf("Test Files/Input1.pdf",db))
