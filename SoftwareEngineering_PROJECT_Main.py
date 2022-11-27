from PyQt5 import QtWidgets as qtw
from PyQt5.QtWidgets import QTableWidgetItem, QMessageBox, QFileDialog
from PyQt5 import QtCore as qtc

#This imports the UI Layout code from the Input, Output, Elective, AutoElective and Settings Python GUI files
from Project_Input_GUI import Ui_Input_window_frm
from Project_Output_GUI import Ui_Output_window_frm
from Project_Elective_GUI import Ui_Elective_popupWindow_frm
from Project_AutoElective_GUI import Ui_AutoElective_popupWindow_frm
from Project_Settings_GUI import Ui_Settings_window_frm
import sys
#   This allows this program to access Excel files
import pandas as pd
import openpyxl
from openpyxl.formula.translate import Translator
#   This allows a File Path check
import os.path as fp

sys.path.insert(1, 'Data') # Makes it so we can access Database.py
import Database
db = Database.Database("Data/database.txt")
import Student # Import Brandon's Student Class to be able to call his Python program
import excelwriter

# Gets access to AI Module
sys.path.insert(1, 'Data') # Makes it so we can access AI_main.py
import AI_main

# Import Eriq's Finalized EXCEL document so it can be displayed in UI
#sys.path.insert(1, '..') # Makes it so we can access Eriq's Excel File
excelOutputFilePath = "" # Path to the output file
excelTemplateFilePath = "Data/Path To Graduation Template.xlsx" # Template file to base outputs on

# Global variable for storing Settings, Manually Chosen ELectives/Hours, and Chosen Degree Track
settings = {"Set_Credits":False, "Manual Elective":False, "Major":""}
finalElectiveList = []

#Sets the Scaling for the UI and all the child widgets to be correct, no matter the screen resolution
qtw.QApplication.setAttribute(qtc.Qt.AA_EnableHighDpiScaling, True)
qtw.QApplication.setAttribute(qtc.Qt.AA_UseHighDpiPixmaps, True)


#---------------------------------------------------------------------------------
#   This creates a CLASS that utlizes the Input_GUI python file to contruct the UI
class InputWindow(qtw.QWidget):

    def __init__(self, *args, ** kwargs):
        super().__init__(*args, **kwargs)

        self.ui = Ui_Input_window_frm()
        self.ui.setupUi(self)
        self.setFixedSize(655,492)
        self.show()

        #
        #----------- Input UI Widget ACTION Definitions ---------------
        #

        #    When the Continue button is clicked, the program will store the
        #--- text field input as a String called 'inputFilePath'
        self.ui.Continue_btn.clicked.connect(self.storeInputForm)

        #    When the BROWSE button is clicked, the program will allow
        #--- the user to browse for the filename and update the Input text_Edit field
        self.ui.Browse_input_file_btn.clicked.connect(self.browseInputFile)

        #    Then the Clear button is clicked, the program will clear
        #--- the lineEdit field for the InputFilePath_lineEdit
        self.ui.Clear_btn.clicked.connect(self.clearInputForm)

        #    When the Settings button is clicked, the program will open
        #--- the 'Settings' Window
        self.ui.Settings_btn.clicked.connect(self.showSettingsWindow)

        #    When the Exit Scheduler button is clicked, the program will close
        self.ui.Exit_application_btn.clicked.connect(self.closeApplication)



    #
    #----------- Input UI Widget Methods/Functions Definitions ---------------
    #

    #   This will store the 'String' for Input File Location
    #---This defines the functionality of the 'Continue_btn'
    def storeInputForm(self):

        global inputStudentID, inputFilePath
        # Making these global so this function can modify them

        # Stores Input for Student ID and Excel Output File Path
        inputStudentID = self.ui.InputStudentID_lineEdit.text()
        inputFilePath = self.ui.InputFilePath_lineEdit.text()

        # Checks to see what Degree Track was Chosen
        settings["Major"] = self.ui.ChooseDegree_cbx.currentText()


        # Checks if File Path exists and runs program if TRUE
        filePath_exists = fp.exists(inputFilePath)
        if filePath_exists:
            # If Manual Selection is TRUE or FALSE, this will display the correct window
            if settings["Manual Elective"]:
                electiveWindow.show()
            else:
                autoElectiveWindow.show()
            #outputWindow.show()

            inputWindow.hide()
            # This clears the Student ID box and filepath box
            self.ui.InputStudentID_lineEdit.clear()
            self.ui.InputFilePath_lineEdit.clear()
        else:
            # Displays Error Message
            qtw.QMessageBox.warning(self, "Error", "File Not Found")

    #   This defines the functionality for the "Clear_btn"
    def clearInputForm(self):
        self.ui.InputStudentID_lineEdit.clear()
        self.ui.InputFilePath_lineEdit.clear()

    #  This defines the functionality for the 'Settings_btn'
    def showSettingsWindow(self):
        inputWindow.hide()
        settingsWindow.show()

    #  This defines the functionality for the 'Exit_scheduler_btn'
    def closeApplication(self):
        quit()

    def browseInputFile(self):
        browseInputFileName = QFileDialog.getOpenFileName(self, 'Choose file', 'C:')
        self.ui.InputFilePath_lineEdit.setText(browseInputFileName[0])




#----------------------------------------------------------------------------------
#   This creates a CLASS that utlizes the Output_GUI python file to contruct the UI
class OutputWindow(qtw.QWidget):

    def __init__(self, *args, ** kwargs):
        super().__init__(*args, **kwargs)

        self.ui = Ui_Output_window_frm()
        self.ui.setupUi(self)
        self.setFixedSize(655,492)

        #
        #----------- Output UI Widget ACTION Definitions ---------------
        #

        # When the Generate Button is clicked, the program will pull the Excel
        # --- file and display the data in the QTable in the UI
        self.ui.GenerateSchedule_btn.clicked.connect(self.generateSchedule)
        #self.ui.GenerateSchedule_btn.clicked.connect(lambda _, xL_Path=excelOutputFilePath, sheet_Name='Sheet1': self.generateSchedule(xL_Path, sheet_Name))

        # When Return Button is clicked, the user will return to Main Menu Window
        self.ui.Return_btn.clicked.connect(self.returnToMainMenu)

        #    When the BROWSE Output button is clicked, the program will allow
        #--- the user to browse for the filename and update the Output text_Edit field
        self.ui.Browse_output_filename_btn.clicked.connect(self.browseOutputFile)

        #When check for prerequisites is clicked run the check schedule function
        self.ui.CheckSchedule_btn.clicked.connect(self.checkSchedule)




    #
    #----------- Output UI Widget Methods/Functions Definitions ---------------
    #

    #    This loads the Finalized Excel File into the Outout GUI QTable
    def generateSchedule(self):
    #def generateSchedule(self, excel_file_path , excel_file_name):

        ew=excelwriter.ExcelWriter()
        st =  Student.Student(db)
        if settings["Set_Credits"]:
            schedule_template = settings["Credits"]
        else:
            schedule_template = {
            "Fall": 15,
            "Spring": 15,
            "Summer": 3
            }
        schedule, classlist = st.generate_schedule(inputFilePath, schedule_template)
        excelOutputFilePath = f"Path to Graduation {inputStudentID}.xlsx"
        errorcheck = ew.savetofile(schedule,classlist,excelOutputFilePath,excelTemplateFilePath ) # Not sure about what the output filename should be



        excelWorkbook = openpyxl.load_workbook(excelOutputFilePath)
        excelWorkSheet = excelWorkbook.active
        # Also, the'Browse' for Output Button is now ENABLED
        self.ui.Browse_output_filename_btn.setEnabled(True)
        # This re-enables the 'Open In Excel' Button
        self.ui.OpenExcelFile_btn.setEnabled(True)

        # Sets Yearly Credit Totals for Spreadsheet in Output UI Table
        # --- Note: This is a workaround because python will not calculate
        # --- Excel formulas

        totalCreditHours = 0
        # --- Year A
        fallYearA_columnTotal = 0
        for i in range(4,11):
            if excelWorkSheet.cell(row=i, column=2).value:
                fallYearA_columnTotal += excelWorkSheet.cell(row=i, column=2).value
            excelWorkSheet.cell(row=11, column=2).value = fallYearA_columnTotal
        totalCreditHours += fallYearA_columnTotal

        springYearA_columnTotal = 0
        for i in range(4,11):
            if excelWorkSheet.cell(row=i, column=4).value:
                springYearA_columnTotal += excelWorkSheet.cell(row=i, column=4).value
            excelWorkSheet.cell(row=11, column=4).value = springYearA_columnTotal
        totalCreditHours += springYearA_columnTotal

        summerYearA_columnTotal = 0
        for i in range(4,6):
            if excelWorkSheet.cell(row=i, column=6).value:
                summerYearA_columnTotal += excelWorkSheet.cell(row=i, column=6).value
            excelWorkSheet.cell(row=6, column=6).value = summerYearA_columnTotal
        totalCreditHours += summerYearA_columnTotal

        #--- Year B
        fallYearB_columnTotal= 0
        for i in range(13,20):
            if excelWorkSheet.cell(row=i, column=2).value:
                fallYearB_columnTotal += excelWorkSheet.cell(row=i, column=2).value
            excelWorkSheet.cell(row=20, column=2).value = fallYearB_columnTotal
        totalCreditHours += fallYearB_columnTotal

        springYearB_columnTotal= 0
        for i in range(13,20):
            if excelWorkSheet.cell(row=i, column=4).value:
                springYearB_columnTotal += excelWorkSheet.cell(row=i, column=4).value
            excelWorkSheet.cell(row=20, column=4).value = springYearB_columnTotal
        totalCreditHours += springYearB_columnTotal

        summerYearB_columnTotal = 0
        for i in range(13,15):
            if excelWorkSheet.cell(row=i, column=6).value:
                summerYearB_columnTotal += excelWorkSheet.cell(row=i, column=6).value
            excelWorkSheet.cell(row=15, column=6).value = summerYearB_columnTotal
        totalCreditHours += summerYearB_columnTotal

        #--- Year C
        fallYearC_columnTotal= 0
        for i in range(22,29):
            if excelWorkSheet.cell(row=i, column=2).value:
                fallYearC_columnTotal += excelWorkSheet.cell(row=i, column=2).value
            excelWorkSheet.cell(row=29, column=2).value = fallYearC_columnTotal
        totalCreditHours += fallYearC_columnTotal

        springYearC_columnTotal= 0
        for i in range(22,29):
            if excelWorkSheet.cell(row=i, column=4).value:
                springYearC_columnTotal += excelWorkSheet.cell(row=i, column=4).value
            excelWorkSheet.cell(row=29, column=4).value = springYearC_columnTotal
        totalCreditHours += springYearC_columnTotal

        summerYearC_columnTotal = 0
        for i in range(22,24):
            if excelWorkSheet.cell(row=i, column=6).value:
                summerYearC_columnTotal += excelWorkSheet.cell(row=i, column=6).value
            excelWorkSheet.cell(row=24, column=6).value = summerYearC_columnTotal
        totalCreditHours += summerYearC_columnTotal

        #--- Year D
        fallYearD_columnTotal= 0
        for i in range(31,38):
            if excelWorkSheet.cell(row=i, column=2).value:
                fallYearD_columnTotal += excelWorkSheet.cell(row=i, column=2).value
            excelWorkSheet.cell(row=38, column=2).value = fallYearD_columnTotal
        totalCreditHours += fallYearD_columnTotal

        springYearD_columnTotal= 0
        for i in range(31,38):
            if excelWorkSheet.cell(row=i, column=4).value:
                springYearD_columnTotal += excelWorkSheet.cell(row=i, column=4).value
            excelWorkSheet.cell(row=38, column=4).value = springYearD_columnTotal
        totalCreditHours += springYearD_columnTotal

        summerYearD_columnTotal = 0
        for i in range(31,33):
            if excelWorkSheet.cell(row=i, column=6).value:
                summerYearD_columnTotal += excelWorkSheet.cell(row=i, column=6).value
            excelWorkSheet.cell(row=33, column=6).value = summerYearD_columnTotal
        totalCreditHours += summerYearD_columnTotal

        # Years A through D Total Credit hours
        excelWorkSheet.cell(row=38, column=6).value = totalCreditHours




        # This portion of code sets the Rows and Columns, also removes NULL cells
        self.ui.SchedulerOutput_tbl.setRowCount(excelWorkSheet .max_row)
        self.ui.SchedulerOutput_tbl.setColumnCount(excelWorkSheet .max_column)

        excelWorkSheet_values_list = list(excelWorkSheet.values)
        self.ui.SchedulerOutput_tbl.setHorizontalHeaderLabels(excelWorkSheet_values_list[0])


        # This portion of code goes through the Excel File and Pulls the Data and Sets Data
        row_Index = 0
        for excelWorkSheet_row in excelWorkSheet_values_list:
            column_Index = 0
            for cell_value in excelWorkSheet_row :
                if cell_value == None:
                    cell_value = ''
                self.ui.SchedulerOutput_tbl.setItem(row_Index, column_Index, QTableWidgetItem(str(cell_value)))
                column_Index += 1
            row_Index += 1

#------------------------------------------------------

    #  This handles the 'Return' button functionality
    def returnToMainMenu(self):
        outputWindow.hide()
        inputWindow.show()

    def checkSchedule(self):
        # Checks if File Path exists and runs program if TRUE
        outputFilePath = self.ui.OutputFilePath_lineEdit.text()
        filePath_exists = fp.exists(outputFilePath)
        st =  Student.Student(db)
        if filePath_exists:
            check = st.check_schedule(outputFilePath)[0]
            remaining = st.check_schedule(outputFilePath)[1]
            #if there is no error say there are no errors. else print the errors that were found
            if check == True:
                qtw.QMessageBox.warning(self, "No prerequisite errors", "No errors were in the excel sheet")
            else:
                cl ="Classes had prerequisite errors:"
                for classes in remaining:
                    cl = cl+ "\n" + classes
                qtw.QMessageBox.warning(self, "Prerequisite errors found", cl)


    #  This handles the 'Browse' output button functionality
    def browseOutputFile(self):
        browseOutputFileName = QFileDialog.getOpenFileName(self, 'Choose file', 'C:')
        self.ui.OutputFilePath_lineEdit.setText(browseOutputFileName[0])
        # Also, 'Check Prerequisite' Button and is now ENABLED
        self.ui.CheckSchedule_btn.setEnabled(True)




#----------------------------------------------------------------------------------
#   This creates a CLASS that utlizes the Settings_GUI python file to contruct the UI
class SettingsWindow(qtw.QWidget):

    def __init__(self, *args, ** kwargs):
        super().__init__(*args, **kwargs)

        self.ui = Ui_Settings_window_frm()
        self.ui.setupUi(self)
        self.setFixedSize(332,246)


    #
    #----------- SETTINGS UI Widget ACTION Definitions ---------------
    #

        # When the Exit Button is clicked, the program will exit the
        # --- Settings Menu and return to Main Input Menu
        self.ui.Exit_btn.clicked.connect(self.exitToMainMenu)

    #
    #----------- SETTINGS UI Widget Methods/Functions Definitions ---------------
    #

    def get_values(self):
        """Return the Settings as a dictionary"""
        try:
            return {
            "Set_Credits" : self.ui.Set_credit_limits_checkBox.isChecked(),
            "Manual Elective": self.ui.Manual_selection_checkBox.isChecked(),
            "Credits" : {
                "Fall" : int(self.ui.lineEdit.text()),
                "Spring": int(self.ui.lineEdit_2.text()),
                "Summer": int(self.ui.lineEdit_3.text())
            }
            }
        except ValueError:
            return {
                "Set_Credits" : self.ui.Set_credit_limits_checkBox.isChecked(),
                "Manual Elective": self.ui.Manual_selection_checkBox.isChecked(),
                "Credits" : {
                    "Fall" : 0,
                    "Spring": 0,
                    "Summer": 0
                }
                }

    def validate_values(self):
        """Check if the settings menu has valid values"""
        credit_hours = [self.ui.lineEdit.text(),self.ui.lineEdit_2.text(),self.ui.lineEdit_3.text()]
        for hour in credit_hours:
            if not hour.isnumeric():
                return False
        return True

    #  This defines the functionality for the 'Exit_btn'
    def exitToMainMenu(self):
        global settings
        settings = self.get_values()
        if self.validate_values() or not settings["Set_Credits"]:
            settingsWindow.hide()
            inputWindow.show()
        else:
            settings["Manual Elective"]:False
            qtw.QMessageBox.warning(self, "Error", "Invalid Settings Entered")
        #settingsWindow.hide()
        #inputWindow.show()



#----------------------------------------------------------------------------------
#   This creates a CLASS that utlizes the Elective_GUI python file to contruct the UI
class ElectiveWindow(qtw.QWidget):

    def __init__(self, *args, ** kwargs):
        super().__init__(*args, **kwargs)

        self.ui = Ui_Elective_popupWindow_frm()
        self.ui.setupUi(self)
        self.setFixedSize(655,492)



    #
    #----------- ELECTIVE UI Widget ACTION Definitions ---------------
    #
        #     When 'Start Selection' Button is clicked, the Electives will be made
        # --- avialble to the user to choose from
        self.ui.StartSelection_btn.clicked.connect(self.startSelection)

        #     When the 'Add Elective' button is clicked, the selected elective
        # --- will be added to the 2nd Combobox
        # --- Once the 2nd Combobox, has 3 electives the button is disabled
        self.ui.AddElective_btn.clicked.connect(self.addElective)

        #     When the 'Continue' Button is clicked, the Elective window will be hidden
        # --- and the Output Window will be displayed
        self.ui.Continue_btn.clicked.connect(self.continueToOutput)

    #
    #----------- ELECTIVE UI Widget Methods/Functions Definitions ---------------
    #

    # This defines the functionality of the 'Add Elective' Button
    # --- This stops after 3 electives are added and 'returns' the final Elective list
    def addElective(self):
        if  self.ui.ShowChosenElectives_listbox.count() < 4:
            self.ui.AddElective_btn.setEnabled(True)
            selectedItem = self.ui.ChooseElective_cbx.currentText()
            self.ui.ShowChosenElectives_listbox.addItem(selectedItem)
            removedItemIndex = self.ui.ChooseElective_cbx.currentIndex()
            self.ui.ChooseElective_cbx.removeItem(removedItemIndex)
            finalElectiveList.append(selectedItem)
            if self.ui.ShowChosenElectives_listbox.count() == 3:
                self.ui.AddElective_btn.setEnabled(False)
                self.ui.Continue_btn.setEnabled(True)
                # Adds Electives to Dynamic AI Databse for AI_main.py module <<<---------------
                for addedElective in finalElectiveList:
                    AI_main.saveToDynamicDatabase(settings["Major"], addedElective)


    # This defines the functionality of the 'Continue' button
    def continueToOutput(self):
        electiveWindow.hide()
        outputWindow.show()
        # This clears the ListBox containing the previous user's choices for electives
        self.ui.ShowChosenElectives_listbox.clear()
        self.ui.StartSelection_btn.setEnabled(True)

    # This defines the functionality of the 'Start Selection' Button
    def startSelection(self):

        self.ui.AddElective_btn.setEnabled(True)
        #Adds List of Available Electives at CSU
        defaultElectivesFile = open("Data/DefaultElectives_Database.txt", "r")
        defaultElectiveData = defaultElectivesFile.read()
        defaultElectiveList = defaultElectiveData.split("\n")
        defaultElectivesFile.close()

        self.ui.ChooseElective_cbx.clear()
        self.ui.ChooseElective_cbx.addItems(defaultElectiveList)
        self.ui.StartSelection_btn.setEnabled(False)
        self.ui.AddElective_btn.setEnabled(True)
        defaultElectiveList.clear()


#----------------------------------------------------------------------------------
#   This creates a CLASS that utlizes the AutoElective_GUI python file to contruct the UI
class AutoElectiveWindow(qtw.QWidget):

    def __init__(self, *args, ** kwargs):
        super().__init__(*args, **kwargs)

        self.ui = Ui_AutoElective_popupWindow_frm()
        self.ui.setupUi(self)
        self.setFixedSize(655,492)

    #
    #----------- AUTOELECTIVE UI Widget ACTION Definitions ---------------
    #

    # When 'Show Electives' Button is clicked, user will see A.I. generated Elective choices
        self.ui.AutoElectives_Show_btn.clicked.connect(self.displayElectiveChosenByAI)
    # When 'Continue' Button is clicked, user will see the Output Window with updated Electives
        self.ui.AutoElectives_Continue_btn.clicked.connect(self.showOutputWindowFromAutoElectives)

    #
    #----------- AUTOELECTIVE UI Widget Methods/Functions Definitions ---------------
    #

    # Defines the use of A.I. module to fill AUTO ELECTIVE combobox with courses choosen by A.I.
    def displayElectiveChosenByAI(self):

        # This calls the AI_Module method for auto-selecting from the DynamicDatabase (chooses 3 electives)
        aiSelectedList = AI_main.chooseElectives(settings["Major"], 3)
        # This saves the AI Selected List to the Dynamic Database
        for aiSelectedElective in aiSelectedList:
            AI_main.saveToDynamicDatabase(settings["Major"], aiSelectedElective)

        #autoElectiveTestList = ["Elective1", "Elective2", "Elective3"]
        self.ui.AutoElective_listBox.clear()
        self.ui.AutoElective_listBox.addItems(aiSelectedList)
        # Enables 'Continue' Button  and disables 'Show Electives' Button
        self.ui.AutoElectives_Continue_btn.setEnabled(True)
        self.ui.AutoElectives_Show_btn.setEnabled(False)

    # Defines use of Continue button functionality
    def showOutputWindowFromAutoElectives(self):
        outputWindow.show()
        autoElectiveWindow.hide()


# This is the MAIN Application
# --- This creates the INPUT, OUTPUT, ELECTIVE, and SETTINGS GUI CLASS objects
# --- This allows the program to run correctly
if __name__ == '__main__':
    app = qtw.QApplication([])

    inputWindow = InputWindow()
    #inputWindow.show()
    outputWindow = OutputWindow()
    #outputWindow.show()
    settingsWindow = SettingsWindow()
    #settingsWindow.show()
    electiveWindow = ElectiveWindow()
    #electiveWindow.show()
    autoElectiveWindow = AutoElectiveWindow()
    #autoElectiveWindow.show()
    app.exec_()
